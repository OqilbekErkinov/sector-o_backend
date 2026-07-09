from datetime import date, timedelta

from django.utils import timezone
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated

from .models import WorkoutLog, WorkoutExerciseEntry, NutritionLog, UserProgram
from .serializers import WorkoutLogSerializer, NutritionLogSerializer


def _resolve_date_param(request, param_name):
    param = request.query_params.get(param_name)
    if param:
        try:
            return date.fromisoformat(param)
        except ValueError:
            pass
    return None


def _resolve_end_date(request):
    # The client's local "today" may differ from the server's (server runs
    # in UTC — see TIME_ZONE in settings.py), which would otherwise make a
    # just-saved log fall outside a "recent days" window near midnight.
    # Trust the client-supplied end_date when present so date-range windows
    # always line up with whatever date the client used to save entries.
    # This value is also the *navigable* anchor for browsing past weeks —
    # see _resolve_today for the separate "real, current week" anchor.
    return _resolve_date_param(request, 'end_date') or timezone.now().date()


def _resolve_today(request):
    """The client's real 'today', for calendar-week-anchored features (program
    adherence) that must always reflect the current week even while the
    caller is browsing a *different* week's summary via end_date."""
    return _resolve_date_param(request, 'today') or timezone.now().date()


def _monday_of(d):
    """Monday of the calendar (ISO) week containing d."""
    return d - timedelta(days=d.weekday())


def _period_range(request):
    period = request.query_params.get('period', 'week')
    end_date_input = _resolve_end_date(request)
    if period == 'month':
        return 'month', end_date_input - timedelta(days=29), end_date_input
    # 'week' is always the full calendar (Mon–Sun) week containing
    # end_date_input — not a rolling trailing window — so "this week" means
    # the same thing here as it already does for program adherence and the
    # 1RM comparison. Navigating to a past week just changes end_date.
    week_start = _monday_of(end_date_input)
    week_end = week_start + timedelta(days=6)
    return 'week', week_start, week_end


class WorkoutLogListCreateView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = WorkoutLog.objects.filter(user=request.user).prefetch_related('exercises__sets')
        date = request.query_params.get('date')
        if date:
            qs = qs.filter(date=date)
        else:
            qs = qs[:60]
        return Response(WorkoutLogSerializer(qs, many=True).data)

    def post(self, request):
        serializer = WorkoutLogSerializer(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        serializer.save(user=request.user)
        return Response(serializer.data, status=201)


class WorkoutLogDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def get_object(self, request, pk):
        return WorkoutLog.objects.filter(pk=pk, user=request.user).first()

    def patch(self, request, pk):
        instance = self.get_object(request, pk)
        if not instance:
            return Response({'error': 'Not found'}, status=404)
        serializer = WorkoutLogSerializer(instance, data=request.data, partial=True, context={'request': request})
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    def delete(self, request, pk):
        instance = self.get_object(request, pk)
        if not instance:
            return Response({'error': 'Not found'}, status=404)
        instance.delete()
        return Response(status=204)


class WorkoutSummaryView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        period, start_date, end_date = _period_range(request)
        logs = WorkoutLog.objects.filter(
            user=request.user, date__gte=start_date, date__lte=end_date
        ).prefetch_related('exercises__sets')

        sets_by_muscle = {key: 0 for key, _ in WorkoutExerciseEntry.MUSCLE_GROUP_CHOICES}
        workout_days = set()

        for log in logs:
            has_content = False
            for exercise in log.exercises.all():
                set_count = exercise.sets.count()
                if set_count == 0:
                    continue
                has_content = True
                if exercise.muscle_group in sets_by_muscle:
                    sets_by_muscle[exercise.muscle_group] += set_count
            if has_content:
                workout_days.add(log.date.isoformat())

        response_data = {
            'period': period,
            'start_date': start_date.isoformat(),
            'end_date': end_date.isoformat(),
            'total_workout_days': len(workout_days),
            'sets_by_muscle': sets_by_muscle,
        }

        # Deliberately uses _resolve_today, NOT the (possibly navigated)
        # end_date above — adherence answers "am I on track *this* week",
        # which must stay pinned to the real current week even while the
        # caller is browsing the muscle-group chart for a past week.
        adherence = self._program_adherence(request, _resolve_today(request))
        if adherence:
            response_data['program_adherence'] = adherence

        return Response(response_data)

    @staticmethod
    def _program_adherence(request, end_date):
        active = UserProgram.objects.filter(user=request.user, is_active=True).prefetch_related('days').first()
        if not active:
            return None
        target_days = active.days.count()
        if target_days == 0:
            return None

        # Always the current calendar week, independent of the period=week/
        # month toggle above — "did I hit my program this week" is inherently
        # a calendar-week question, same convention as the 1RM comparison.
        week_start = _monday_of(end_date)
        week_end = week_start + timedelta(days=6)
        completed_days = WorkoutLog.objects.filter(
            user=request.user,
            date__gte=week_start, date__lte=week_end,
            program_day__program=active,
        ).values('program_day').distinct().count()

        return {
            'program_name': active.name,
            'completed_days': completed_days,
            'target_days': target_days,
            'week_start': week_start.isoformat(),
            'week_end': week_end.isoformat(),
        }


class WorkoutComparisonView(APIView):
    """This-week vs last-week strength comparison, matched by (program day,
    exercise name, set number) — e.g. 'Biceps Curl' set 1 on *this week's
    Day 1* vs set 1 on *last week's Day 1*. Scoping by program day matters
    for a multi-day split: if Day 1 and Day 3 happen to share an exercise
    name, they must never be compared against each other — a split is
    followed for months, so each day's own trend has to stay pure. Logs
    with no program day (no active program used) fall back to matching by
    name alone, same as before.
    Only exercises logged in BOTH weeks (within the same day-scope) are
    returned — there's nothing to compare a brand-new exercise against."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        end_date = _resolve_end_date(request)
        # Calendar (Mon–Sun) weeks, not a rolling "last 7 days" window — a
        # rolling window doesn't align with how people think of "this week"
        # vs "last week" and would misclassify e.g. Thursday's log as part
        # of "this week" even after the calendar week has turned over.
        this_week_start = _monday_of(end_date)
        this_week_end = this_week_start + timedelta(days=6)
        last_week_end = this_week_start - timedelta(days=1)
        last_week_start = last_week_end - timedelta(days=6)

        this_week_logs = WorkoutLog.objects.filter(
            user=request.user, date__gte=this_week_start, date__lte=this_week_end
        ).select_related('program_day').prefetch_related('exercises__sets').order_by('date')
        last_week_logs = WorkoutLog.objects.filter(
            user=request.user, date__gte=last_week_start, date__lte=last_week_end
        ).select_related('program_day').prefetch_related('exercises__sets').order_by('date')

        this_week_map = self._latest_exercise_by_day_and_name(this_week_logs)
        last_week_map = self._latest_exercise_by_day_and_name(last_week_logs)

        comparisons = []
        for key, current in this_week_map.items():
            previous = last_week_map.get(key)
            if not previous:
                continue

            prev_by_set_number = {s.set_number: s for s in previous['exercise'].sets.all()}
            set_comparisons = []
            for s in current['exercise'].sets.all():
                prev_set = prev_by_set_number.get(s.set_number)
                set_comparisons.append({
                    'set_number': s.set_number,
                    'current_weight_kg': s.weight_kg,
                    'current_reps': s.reps,
                    'previous_weight_kg': prev_set.weight_kg if prev_set else None,
                    'previous_reps': prev_set.reps if prev_set else None,
                    **self._compute_changes(s, prev_set),
                })

            comparisons.append({
                'name': current['exercise'].name,
                'muscle_group': current['exercise'].muscle_group,
                'program_day_name': current['program_day_name'],
                'current_date': current['date'].isoformat(),
                'previous_date': previous['date'].isoformat(),
                'sets': set_comparisons,
            })

        return Response({
            'this_week_start': this_week_start.isoformat(),
            'this_week_end': this_week_end.isoformat(),
            'last_week_start': last_week_start.isoformat(),
            'last_week_end': last_week_end.isoformat(),
            'exercises': comparisons,
        })

    @staticmethod
    def _latest_exercise_by_day_and_name(logs_ascending):
        """logs_ascending must be ordered by date ascending, so the last
        write for a given (program day, exercise name) is that week's most
        recent one. Logs without a program_day key on None, which still
        keeps them isolated from any actual program day's exercises."""
        result = {}
        for log in logs_ascending:
            for exercise in log.exercises.all():
                key = (log.program_day_id, exercise.name.strip().lower())
                result[key] = {
                    'date': log.date,
                    'exercise': exercise,
                    'program_day_name': log.program_day.name if log.program_day_id else None,
                }
        return result

    @staticmethod
    def _estimated_1rm(weight_kg, reps):
        """Epley formula: estimates the one-rep max implied by a given
        weight×reps set. Weight and reps alone can't be compared directly
        (heavier-for-fewer vs lighter-for-more aren't equivalent effort) —
        1RM converts any set to the same "true max strength" unit so two
        different weight/rep combinations become one comparable number."""
        return weight_kg * (1 + reps / 30)

    def _compute_changes(self, current_set, prev_set):
        current_1rm = round(self._estimated_1rm(current_set.weight_kg, current_set.reps), 1)
        if not prev_set:
            return {
                'current_1rm_kg': current_1rm,
                'previous_1rm_kg': None,
                'change_pct': None,
            }
        previous_1rm = round(self._estimated_1rm(prev_set.weight_kg, prev_set.reps), 1)
        return {
            'current_1rm_kg': current_1rm,
            'previous_1rm_kg': previous_1rm,
            'change_pct': (
                round((current_1rm - previous_1rm) / previous_1rm * 100, 1)
                if previous_1rm else None
            ),
        }


class WorkoutExportView(APIView):
    """Exports the workout diary as a formatted .xlsx laid out like a
    classic paper/spreadsheet training log: **rows** are the active
    program's days, each followed immediately by its own exercises (one
    exercise per row, in program order) — a colored header row marks where
    each day's block starts. **Columns** are calendar (Mon–Sun) weeks in
    chronological order (oldest on the left, matching how progress is
    normally read left-to-right), and each week is itself split into one
    sub-column per set ("1-padxod", "2-padxod", ...) so every set gets its
    own cell instead of being crammed together. Exercises logged without
    (or outside) the active program are appended in their own trailing
    block(s), grouped by whatever day label they do have, so no data is
    silently dropped."""
    permission_classes = [IsAuthenticated]

    SCOPE_DAYS = {'day': 0, 'week': 6, 'month': 29}

    def get(self, request):
        from collections import OrderedDict
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse

        scope = request.query_params.get('scope', 'week')
        end_date = _resolve_end_date(request)

        qs = WorkoutLog.objects.filter(user=request.user, date__lte=end_date)
        if scope in self.SCOPE_DAYS:
            qs = qs.filter(date__gte=end_date - timedelta(days=self.SCOPE_DAYS[scope]))
        logs = list(qs.select_related('program_day').prefetch_related('exercises__sets').order_by('date'))

        muscle_labels = dict(WorkoutExerciseEntry.MUSCLE_GROUP_CHOICES)

        # data[(program_day_id_or_None, exercise_name_lower)][week_monday] = exercise
        # "Latest wins" if an exercise is logged more than once in the same
        # week, same convention used everywhere else in this file.
        data = {}
        week_mondays = set()
        max_sets = 1
        for log in logs:
            monday = _monday_of(log.date)
            week_mondays.add(monday)
            for exercise in log.exercises.all():
                key = (log.program_day_id, exercise.name.strip().lower())
                data.setdefault(key, {})[monday] = exercise
                max_sets = max(max_sets, exercise.sets.count())
        sorted_weeks = sorted(week_mondays)  # ascending: oldest week first

        # Row blocks: the active program's days (in their defined order, so
        # every exercise gets a row even in a week it wasn't logged), then
        # any other day-groupings actually found in the data.
        active_program = UserProgram.objects.filter(
            user=request.user, is_active=True
        ).prefetch_related('days__exercises').first()

        row_blocks = []
        covered_keys = set()
        if active_program:
            for day in active_program.days.all():
                rows = [
                    (ex.name, ex.muscle_group, day.id, ex.name.strip().lower())
                    for ex in day.exercises.all()
                ]
                covered_keys.update((day.id, name_lower) for _, _, _, name_lower in rows)
                row_blocks.append((day.name, rows))

        extra_blocks = OrderedDict()
        for log in logs:
            day_label = log.program_day.name if log.program_day_id else 'Erkin mashqlar'
            for exercise in log.exercises.all():
                name_lower = exercise.name.strip().lower()
                key = (log.program_day_id, name_lower)
                if key in covered_keys:
                    continue
                block = extra_blocks.setdefault(day_label, OrderedDict())
                block.setdefault(key, (exercise.name, exercise.muscle_group, log.program_day_id, name_lower))
                covered_keys.add(key)
        for label, rows in extra_blocks.items():
            row_blocks.append((label, list(rows.values())))

        wb = Workbook()
        ws = wb.active
        ws.title = "Trening tarixi"

        header_fill = PatternFill(start_color="00CFFF", end_color="00CFFF", fill_type="solid")
        header_font = Font(bold=True, color="000000")
        day_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
        day_font = Font(bold=True, color="FFFFFF")
        wrap = Alignment(wrap_text=True, vertical="top")
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        last_col = 1 + len(sorted_weeks) * max_sets

        # ── Two-row header: "Mashq" + one merged "Hafta N" block per week,
        # each block subdivided into "1-padxod", "2-padxod", ... columns. ──
        ws.cell(row=1, column=1, value="Mashq").font = header_font
        ws.cell(row=1, column=1).fill = header_fill
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        ws.column_dimensions['A'].width = 30

        for week_idx, monday in enumerate(sorted_weeks):
            sunday = monday + timedelta(days=6)
            start_col = 2 + week_idx * max_sets
            end_col = start_col + max_sets - 1
            week_cell = ws.cell(row=1, column=start_col, value=f"Hafta {week_idx + 1}\n{monday.isoformat()} — {sunday.isoformat()}")
            week_cell.font = header_font
            week_cell.fill = header_fill
            week_cell.alignment = center
            if end_col > start_col:
                ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            for set_idx in range(max_sets):
                col = start_col + set_idx
                set_cell = ws.cell(row=2, column=col, value=f"{set_idx + 1}-padxod")
                set_cell.font = header_font
                set_cell.fill = header_fill
                set_cell.alignment = center
                ws.column_dimensions[get_column_letter(col)].width = 13
                if end_col > start_col:
                    ws.cell(row=1, column=col).fill = header_fill

        row = 3
        for day_label, rows in row_blocks:
            day_cell = ws.cell(row=row, column=1, value=day_label)
            day_cell.font = day_font
            day_cell.fill = day_fill
            if last_col > 1:
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
                for col in range(2, last_col + 1):
                    ws.cell(row=row, column=col).fill = day_fill
            row += 1

            for name, muscle_group, day_key, name_lower in rows:
                muscle = muscle_labels.get(muscle_group, '')
                label = f"{name} ({muscle})" if muscle else name
                ws.cell(row=row, column=1, value=label).alignment = wrap
                for week_idx, monday in enumerate(sorted_weeks):
                    exercise = data.get((day_key, name_lower), {}).get(monday)
                    if not exercise:
                        continue
                    sets = list(exercise.sets.all())
                    start_col = 2 + week_idx * max_sets
                    for set_idx, s in enumerate(sets):
                        cell = ws.cell(row=row, column=start_col + set_idx, value=f"{s.weight_kg}kg×{s.reps}")
                        cell.alignment = center
                row += 1
            row += 1  # spacer row between day blocks

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"trening_{scope}_{end_date.isoformat()}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


class NutritionLogListCreateView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = NutritionLog.objects.filter(user=request.user)
        date = request.query_params.get('date')
        if date:
            qs = qs.filter(date=date)
        else:
            qs = qs[:100]
        return Response(NutritionLogSerializer(qs, many=True).data)

    def post(self, request):
        serializer = NutritionLogSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save(user=request.user)
        return Response(serializer.data, status=201)


class NutritionLogDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def patch(self, request, pk):
        instance = NutritionLog.objects.filter(pk=pk, user=request.user).first()
        if not instance:
            return Response({'error': 'Not found'}, status=404)
        serializer = NutritionLogSerializer(instance, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    def delete(self, request, pk):
        instance = NutritionLog.objects.filter(pk=pk, user=request.user).first()
        if not instance:
            return Response({'error': 'Not found'}, status=404)
        instance.delete()
        return Response(status=204)


class NutritionSummaryView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        period, start_date, end_date = _period_range(request)
        logs = NutritionLog.objects.filter(user=request.user, date__gte=start_date, date__lte=end_date)

        daily = {}
        for log in logs:
            entry = daily.setdefault(
                log.date.isoformat(),
                {'calories': 0, 'protein_g': 0.0, 'carbs_g': 0.0, 'fat_g': 0.0},
            )
            entry['calories'] += log.calories
            entry['protein_g'] += log.protein_g
            entry['carbs_g'] += log.carbs_g
            entry['fat_g'] += log.fat_g

        days = []
        d = start_date
        while d <= end_date:
            key = d.isoformat()
            days.append({'date': key, **daily.get(key, {'calories': 0, 'protein_g': 0.0, 'carbs_g': 0.0, 'fat_g': 0.0})})
            d += timedelta(days=1)

        logged_days = len(daily)
        total_calories = sum(v['calories'] for v in daily.values())

        return Response({
            'period': period,
            'start_date': start_date.isoformat(),
            'end_date': end_date.isoformat(),
            'logged_days': logged_days,
            'total_calories': total_calories,
            'avg_calories_per_logged_day': round(total_calories / logged_days, 1) if logged_days else 0,
            'total_protein_g': sum(v['protein_g'] for v in daily.values()),
            'total_carbs_g': sum(v['carbs_g'] for v in daily.values()),
            'total_fat_g': sum(v['fat_g'] for v in daily.values()),
            'days': days,
        })


class NutritionExportView(APIView):
    """Exports the food diary as a formatted .xlsx: one block per logged
    day (oldest first), each day broken into its meal types (Nonushta,
    Tushlik, Kechki ovqat, ...) in a fixed sensible order, with every food
    entry as its own row — a detailed day-by-day diary printout rather than
    a flattened list, so nothing from the same day gets mixed together."""
    permission_classes = [IsAuthenticated]

    SCOPE_DAYS = {'day': 0, 'week': 6, 'month': 29}
    MEAL_TYPE_ORDER = [choice[0] for choice in NutritionLog.MEAL_TYPE_CHOICES]

    def get(self, request):
        from collections import OrderedDict
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse

        scope = request.query_params.get('scope', 'week')
        end_date = _resolve_end_date(request)

        qs = NutritionLog.objects.filter(user=request.user, date__lte=end_date)
        if scope in self.SCOPE_DAYS:
            qs = qs.filter(date__gte=end_date - timedelta(days=self.SCOPE_DAYS[scope]))
        logs = list(qs.order_by('date', 'created_at'))

        meal_labels = dict(NutritionLog.MEAL_TYPE_CHOICES)

        by_date = OrderedDict()
        for log in logs:
            by_date.setdefault(log.date, []).append(log)

        wb = Workbook()
        ws = wb.active
        ws.title = "Ovqatlanish tarixi"

        header_fill = PatternFill(start_color="00CFFF", end_color="00CFFF", fill_type="solid")
        header_font = Font(bold=True, color="000000")
        date_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
        date_font = Font(bold=True, color="FFFFFF")
        meal_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
        meal_font = Font(bold=True, color="00CFFF")
        wrap = Alignment(wrap_text=True, vertical="top")
        last_col = 5

        headers = ['Ovqat', 'Kaloriya', 'Oqsil (g)', 'Uglevod (g)', "Yog' (g)"]
        for col, label in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=label)
            cell.font = header_font
            cell.fill = header_fill
        ws.column_dimensions['A'].width = 32
        for col in range(2, last_col + 1):
            ws.column_dimensions[get_column_letter(col)].width = 14

        row = 2
        for log_date, day_logs in by_date.items():
            day_calories = sum(l.calories for l in day_logs)
            day_protein = sum(l.protein_g for l in day_logs)
            day_carbs = sum(l.carbs_g for l in day_logs)
            day_fat = sum(l.fat_g for l in day_logs)

            date_cell = ws.cell(
                row=row, column=1,
                value=f"{log_date.isoformat()} — Jami: {day_calories} kal "
                      f"(Oqsil {day_protein:g}g / Uglevod {day_carbs:g}g / Yog' {day_fat:g}g)",
            )
            date_cell.font = date_font
            date_cell.fill = date_fill
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
            for col in range(2, last_col + 1):
                ws.cell(row=row, column=col).fill = date_fill
            row += 1

            by_meal = OrderedDict()
            for log in day_logs:
                by_meal.setdefault(log.meal_type, []).append(log)

            for meal_type in self.MEAL_TYPE_ORDER:
                meal_logs = by_meal.get(meal_type)
                if not meal_logs:
                    continue
                meal_calories = sum(l.calories for l in meal_logs)
                meal_cell = ws.cell(row=row, column=1, value=f"{meal_labels[meal_type]} — {meal_calories} kal")
                meal_cell.font = meal_font
                meal_cell.fill = meal_fill
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
                for col in range(2, last_col + 1):
                    ws.cell(row=row, column=col).fill = meal_fill
                row += 1

                for log in meal_logs:
                    ws.cell(row=row, column=1, value=log.name).alignment = wrap
                    ws.cell(row=row, column=2, value=log.calories)
                    ws.cell(row=row, column=3, value=log.protein_g)
                    ws.cell(row=row, column=4, value=log.carbs_g)
                    ws.cell(row=row, column=5, value=log.fat_g)
                    row += 1
            row += 1  # spacer row between days

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"ovqatlanish_{scope}_{end_date.isoformat()}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
