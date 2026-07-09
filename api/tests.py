from datetime import timedelta

from django.core.cache import cache
from django.urls import reverse
from django.utils import timezone
from rest_framework.test import APITestCase

from api.models import (
    OTPCode, WorkoutLog, WorkoutSet, NutritionLog, UserProgram, UserProgramDay,
    Category, Exercise, Program, ProgramDay,
)

from django.contrib.auth import get_user_model

User = get_user_model()


class BaseAPITestCase(APITestCase):
    """Clears the throttle cache before each test so ScopedRateThrottle state
    from one test doesn't leak into (and randomly fail) another."""

    def setUp(self):
        cache.clear()


class RegisterAndLoginTests(BaseAPITestCase):
    def test_register_creates_inactive_user_and_otp(self):
        resp = self.client.post(reverse('register'), {
            'email': 'test@example.com',
            'password': 'strongpass123',
            'full_name': 'Test User',
        })
        self.assertEqual(resp.status_code, 201)
        user = User.objects.get(email='test@example.com')
        self.assertFalse(user.is_active)
        self.assertTrue(OTPCode.objects.filter(email='test@example.com', purpose='verify').exists())

    def test_verify_email_activates_user_and_returns_tokens(self):
        self.client.post(reverse('register'), {
            'email': 'verify@example.com',
            'password': 'strongpass123',
        })
        otp = OTPCode.objects.filter(email='verify@example.com', purpose='verify').latest('created_at')

        resp = self.client.post(reverse('verify-email'), {
            'email': 'verify@example.com',
            'code': otp.code,
        })
        self.assertEqual(resp.status_code, 200)
        self.assertIn('access', resp.data)
        user = User.objects.get(email='verify@example.com')
        self.assertTrue(user.is_active)

    def test_verify_email_rejects_wrong_code(self):
        self.client.post(reverse('register'), {
            'email': 'wrongcode@example.com',
            'password': 'strongpass123',
        })
        resp = self.client.post(reverse('verify-email'), {
            'email': 'wrongcode@example.com',
            'code': '000000',
        })
        self.assertEqual(resp.status_code, 400)

    def test_login_requires_verified_email(self):
        User.objects.create_user(email='unverified@example.com', password='strongpass123')
        resp = self.client.post(reverse('login'), {
            'email': 'unverified@example.com',
            'password': 'strongpass123',
        })
        self.assertEqual(resp.status_code, 403)
        self.assertTrue(resp.data.get('needs_verification'))

    def test_login_rejects_wrong_password(self):
        user = User.objects.create_user(email='loginuser@example.com', password='strongpass123')
        user.is_active = True
        user.save()
        resp = self.client.post(reverse('login'), {
            'email': 'loginuser@example.com',
            'password': 'wrongpass',
        })
        self.assertEqual(resp.status_code, 401)


class AddXPTests(BaseAPITestCase):
    def setUp(self):
        super().setUp()
        self.user = User.objects.create_user(email='xp@example.com', password='strongpass123')
        self.user.is_active = True
        self.user.save()
        resp = self.client.post(reverse('login'), {'email': 'xp@example.com', 'password': 'strongpass123'})
        self.token = resp.data['access']

    def auth_header(self):
        return {'HTTP_AUTHORIZATION': f'Bearer {self.token}'}

    def test_add_xp_within_limit_succeeds(self):
        resp = self.client.post(reverse('add-xp'), {'xp': 50}, **self.auth_header())
        self.assertEqual(resp.status_code, 200)
        self.user.refresh_from_db()
        self.assertEqual(self.user.xp, 50)

    def test_add_xp_above_cap_is_rejected(self):
        resp = self.client.post(reverse('add-xp'), {'xp': 999999}, **self.auth_header())
        self.assertEqual(resp.status_code, 400)
        self.user.refresh_from_db()
        self.assertEqual(self.user.xp, 0)

    def test_add_xp_rejects_zero_or_negative(self):
        resp = self.client.post(reverse('add-xp'), {'xp': 0}, **self.auth_header())
        self.assertEqual(resp.status_code, 400)


class EmailChangeTests(BaseAPITestCase):
    def setUp(self):
        super().setUp()
        self.user = User.objects.create_user(email='old@example.com', password='strongpass123')
        self.user.is_active = True
        self.user.save()
        resp = self.client.post(reverse('login'), {'email': 'old@example.com', 'password': 'strongpass123'})
        self.token = resp.data['access']

    def auth_header(self):
        return {'HTTP_AUTHORIZATION': f'Bearer {self.token}'}

    def test_patch_email_does_not_change_immediately(self):
        resp = self.client.patch(
            reverse('me'), {'email': 'new@example.com'},
            format='json', **self.auth_header()
        )
        self.assertEqual(resp.status_code, 200)
        self.user.refresh_from_db()
        self.assertEqual(self.user.email, 'old@example.com')
        self.assertTrue(
            OTPCode.objects.filter(email='old@example.com', purpose='change_email', new_email='new@example.com').exists()
        )

    def test_confirm_email_change_with_valid_code_applies_change(self):
        self.client.patch(reverse('me'), {'email': 'confirmed@example.com'}, format='json', **self.auth_header())
        otp = OTPCode.objects.filter(email='old@example.com', purpose='change_email').latest('created_at')

        resp = self.client.post(reverse('confirm-email-change'), {'code': otp.code}, **self.auth_header())
        self.assertEqual(resp.status_code, 200)
        self.user.refresh_from_db()
        self.assertEqual(self.user.email, 'confirmed@example.com')

    def test_confirm_email_change_with_wrong_code_fails(self):
        self.client.patch(reverse('me'), {'email': 'nope@example.com'}, format='json', **self.auth_header())
        resp = self.client.post(reverse('confirm-email-change'), {'code': '000000'}, **self.auth_header())
        self.assertEqual(resp.status_code, 400)
        self.user.refresh_from_db()
        self.assertEqual(self.user.email, 'old@example.com')


class WorkoutTrackingTests(BaseAPITestCase):
    def setUp(self):
        super().setUp()
        self.user = User.objects.create_user(email='lifter@example.com', password='strongpass123')
        self.user.is_active = True
        self.user.save()
        resp = self.client.post(reverse('login'), {'email': 'lifter@example.com', 'password': 'strongpass123'})
        self.token = resp.data['access']

    def auth_header(self):
        return {'HTTP_AUTHORIZATION': f'Bearer {self.token}'}

    def _create_log(self, date=None, muscle_group='biceps'):
        payload = {
            'date': date or timezone.now().date().isoformat(),
            'notes': 'Push day',
            'exercises': [
                {
                    'name': 'Barbell Curl',
                    'muscle_group': muscle_group,
                    'order': 0,
                    'sets': [
                        {'set_number': 1, 'weight_kg': 20, 'reps': 10},
                        {'set_number': 2, 'weight_kg': 22.5, 'reps': 8},
                    ],
                }
            ],
        }
        return self.client.post(reverse('workout-logs'), payload, format='json', **self.auth_header())

    def test_create_workout_log_with_nested_sets(self):
        resp = self._create_log()
        self.assertEqual(resp.status_code, 201)
        log = WorkoutLog.objects.get(user=self.user)
        self.assertEqual(log.exercises.count(), 1)
        self.assertEqual(log.exercises.first().sets.count(), 2)

    def test_list_workout_logs_scoped_to_user(self):
        self._create_log()
        other = User.objects.create_user(email='other@example.com', password='strongpass123')
        WorkoutLog.objects.create(user=other, date='2026-06-01')

        resp = self.client.get(reverse('workout-logs'), **self.auth_header())
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(len(resp.data), 1)

    def test_delete_workout_log(self):
        self._create_log()
        log_id = WorkoutLog.objects.get(user=self.user).id
        resp = self.client.delete(reverse('workout-log-detail', args=[log_id]), **self.auth_header())
        self.assertEqual(resp.status_code, 204)
        self.assertFalse(WorkoutLog.objects.filter(id=log_id).exists())

    def test_cannot_delete_other_users_workout_log(self):
        other = User.objects.create_user(email='other2@example.com', password='strongpass123')
        other_log = WorkoutLog.objects.create(user=other, date='2026-06-01')
        resp = self.client.delete(reverse('workout-log-detail', args=[other_log.id]), **self.auth_header())
        self.assertEqual(resp.status_code, 404)
        self.assertTrue(WorkoutLog.objects.filter(id=other_log.id).exists())

    def test_workout_summary_aggregates_sets_by_muscle(self):
        self._create_log(muscle_group='biceps')
        resp = self.client.get(reverse('workout-summary'), {'period': 'week'}, **self.auth_header())
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.data['sets_by_muscle']['biceps'], 2)
        self.assertEqual(resp.data['sets_by_muscle']['chest'], 0)
        self.assertEqual(resp.data['total_workout_days'], 1)
        self.assertEqual(set(resp.data['sets_by_muscle'].keys()), {
            'chest', 'back', 'shoulders', 'biceps', 'triceps', 'legs', 'abs', 'glutes',
        })

    def test_workout_summary_separates_muscle_groups_across_exercises(self):
        payload = {
            'date': timezone.now().date().isoformat(),
            'notes': '',
            'exercises': [
                {'name': 'Bench Press', 'muscle_group': 'chest', 'order': 0, 'sets': [
                    {'set_number': 1, 'weight_kg': 60, 'reps': 8},
                    {'set_number': 2, 'weight_kg': 60, 'reps': 8},
                    {'set_number': 3, 'weight_kg': 60, 'reps': 8},
                ]},
                {'name': 'Squat', 'muscle_group': 'legs', 'order': 1, 'sets': [
                    {'set_number': 1, 'weight_kg': 80, 'reps': 5},
                ]},
            ],
        }
        self.client.post(reverse('workout-logs'), payload, format='json', **self.auth_header())

        resp = self.client.get(reverse('workout-summary'), {'period': 'week'}, **self.auth_header())
        self.assertEqual(resp.data['sets_by_muscle']['chest'], 3)
        self.assertEqual(resp.data['sets_by_muscle']['legs'], 1)
        self.assertEqual(resp.data['sets_by_muscle']['back'], 0)

    def test_workout_summary_uses_client_end_date_across_timezones(self):
        # "week" is a full calendar (Mon-Sun) week, so a same-week client/
        # server day offset (the original timezone bug) never matters here —
        # what must still be respected is end_date picking a genuinely
        # *different* week. Next Monday is guaranteed to be in the next
        # calendar week regardless of which weekday the suite runs on.
        this_monday = timezone.now().date() - timedelta(days=timezone.now().date().weekday())
        next_week_day = (this_monday + timedelta(days=7)).isoformat()
        self._create_log(date=next_week_day, muscle_group='back')

        resp = self.client.get(reverse('workout-summary'), {'period': 'week'}, **self.auth_header())
        self.assertEqual(resp.data['sets_by_muscle']['back'], 0, "current week should not see next week's log")

        resp = self.client.get(
            reverse('workout-summary'), {'period': 'week', 'end_date': next_week_day}, **self.auth_header()
        )
        self.assertEqual(resp.data['sets_by_muscle']['back'], 2)


class NutritionTrackingTests(BaseAPITestCase):
    def setUp(self):
        super().setUp()
        self.user = User.objects.create_user(email='eater@example.com', password='strongpass123')
        self.user.is_active = True
        self.user.save()
        resp = self.client.post(reverse('login'), {'email': 'eater@example.com', 'password': 'strongpass123'})
        self.token = resp.data['access']

    def auth_header(self):
        return {'HTTP_AUTHORIZATION': f'Bearer {self.token}'}

    def test_create_and_list_nutrition_log(self):
        resp = self.client.post(reverse('nutrition-logs'), {
            'name': 'Chicken & rice',
            'calories': 600,
            'protein_g': 45,
            'carbs_g': 70,
            'fat_g': 12,
        }, **self.auth_header())
        self.assertEqual(resp.status_code, 201)

        resp = self.client.get(reverse('nutrition-logs'), **self.auth_header())
        self.assertEqual(len(resp.data), 1)

    def test_cannot_delete_other_users_nutrition_log(self):
        other = User.objects.create_user(email='other3@example.com', password='strongpass123')
        entry = NutritionLog.objects.create(user=other, name='Salad', calories=200)
        resp = self.client.delete(reverse('nutrition-log-detail', args=[entry.id]), **self.auth_header())
        self.assertEqual(resp.status_code, 404)

    def test_can_edit_own_nutrition_log(self):
        create_resp = self.client.post(reverse('nutrition-logs'), {
            'name': 'Rize', 'calories': 300, 'meal_type': 'lunch',
        }, **self.auth_header())
        entry_id = create_resp.data['id']

        resp = self.client.patch(reverse('nutrition-log-detail', args=[entry_id]), {
            'name': 'Rice', 'calories': 350, 'meal_type': 'dinner',
        }, format='json', **self.auth_header())
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.data['name'], 'Rice')
        self.assertEqual(resp.data['calories'], 350)
        self.assertEqual(resp.data['meal_type'], 'dinner')

    def test_cannot_edit_other_users_nutrition_log(self):
        other = User.objects.create_user(email='other4@example.com', password='strongpass123')
        entry = NutritionLog.objects.create(user=other, name='Salad', calories=200)
        resp = self.client.patch(
            reverse('nutrition-log-detail', args=[entry.id]), {'calories': 999}, format='json', **self.auth_header()
        )
        self.assertEqual(resp.status_code, 404)
        entry.refresh_from_db()
        self.assertEqual(entry.calories, 200)

    def test_nutrition_summary_aggregates_macros(self):
        self.client.post(reverse('nutrition-logs'), {
            'name': 'Breakfast', 'calories': 400, 'protein_g': 30, 'carbs_g': 40, 'fat_g': 10,
        }, **self.auth_header())
        self.client.post(reverse('nutrition-logs'), {
            'name': 'Lunch', 'calories': 700, 'protein_g': 50, 'carbs_g': 60, 'fat_g': 20,
        }, **self.auth_header())

        resp = self.client.get(reverse('nutrition-summary'), {'period': 'week'}, **self.auth_header())
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.data['total_calories'], 1100)
        self.assertEqual(resp.data['logged_days'], 1)
        self.assertEqual(resp.data['avg_calories_per_logged_day'], 1100)
        self.assertEqual(len(resp.data['days']), 7)

    def test_nutrition_summary_uses_client_end_date_across_timezones(self):
        # "week" is a full calendar (Mon-Sun) week, so next Monday is
        # guaranteed to be outside the current week regardless of which
        # weekday the suite happens to run on.
        this_monday = timezone.now().date() - timedelta(days=timezone.now().date().weekday())
        next_week_day = (this_monday + timedelta(days=7)).isoformat()
        NutritionLog.objects.create(user=self.user, date=next_week_day, name='Late snack', calories=300)

        resp = self.client.get(reverse('nutrition-summary'), {'period': 'week'}, **self.auth_header())
        self.assertEqual(resp.data['total_calories'], 0)

        resp = self.client.get(
            reverse('nutrition-summary'), {'period': 'week', 'end_date': next_week_day}, **self.auth_header()
        )
        self.assertEqual(resp.data['total_calories'], 300)

    def test_nutrition_log_defaults_to_other_meal_type(self):
        resp = self.client.post(reverse('nutrition-logs'), {
            'name': 'Mystery snack', 'calories': 150,
        }, **self.auth_header())
        self.assertEqual(resp.data['meal_type'], 'other')

    def test_nutrition_log_accepts_meal_type(self):
        resp = self.client.post(reverse('nutrition-logs'), {
            'name': 'Oatmeal', 'calories': 300, 'meal_type': 'breakfast',
        }, **self.auth_header())
        self.assertEqual(resp.status_code, 201)
        self.assertEqual(resp.data['meal_type'], 'breakfast')

    def test_nutrition_export_groups_entries_by_date_and_meal_type(self):
        today = timezone.now().date()
        self.client.post(reverse('nutrition-logs'), {
            'name': 'Oatmeal', 'calories': 300, 'protein_g': 10, 'carbs_g': 50, 'fat_g': 5,
            'meal_type': 'breakfast', 'date': today.isoformat(),
        }, **self.auth_header())
        self.client.post(reverse('nutrition-logs'), {
            'name': 'Chicken Rice', 'calories': 600, 'protein_g': 45, 'carbs_g': 70, 'fat_g': 12,
            'meal_type': 'lunch', 'date': today.isoformat(),
        }, **self.auth_header())

        resp = self.client.get(
            reverse('nutrition-export'), {'scope': 'week', 'end_date': today.isoformat()}, **self.auth_header()
        )
        self.assertEqual(resp.status_code, 200)

        from io import BytesIO
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(resp.content))
        ws = wb.active
        self.assertEqual(ws.cell(row=1, column=1).value, 'Ovqat')
        self.assertIn(today.isoformat(), ws.cell(row=2, column=1).value)
        # Meal types appear in a fixed, sensible order (breakfast before
        # lunch) regardless of the order the entries were created in, and
        # each gets its own sub-header row before its entries.
        self.assertTrue(ws.cell(row=3, column=1).value.startswith('Nonushta'))
        self.assertEqual(ws.cell(row=4, column=1).value, 'Oatmeal')
        self.assertEqual(ws.cell(row=4, column=2).value, 300)
        self.assertTrue(ws.cell(row=5, column=1).value.startswith('Tushlik'))
        self.assertEqual(ws.cell(row=6, column=1).value, 'Chicken Rice')

    def test_updating_nutrition_goals_via_profile_patch(self):
        resp = self.client.patch(reverse('me'), {
            'goal_calories': 2200, 'goal_protein_g': 150, 'goal_carbs_g': 250, 'goal_fat_g': 70,
        }, format='json', **self.auth_header())
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.data['goal_calories'], 2200)
        self.assertEqual(resp.data['goal_protein_g'], 150)
        self.assertEqual(resp.data['goal_carbs_g'], 250)
        self.assertEqual(resp.data['goal_fat_g'], 70)


class WorkoutComparisonAndExportTests(BaseAPITestCase):
    def setUp(self):
        super().setUp()
        self.user = User.objects.create_user(email='progress@example.com', password='strongpass123')
        self.user.is_active = True
        self.user.save()
        resp = self.client.post(reverse('login'), {'email': 'progress@example.com', 'password': 'strongpass123'})
        self.token = resp.data['access']
        self.today = timezone.now().date()

    def auth_header(self):
        return {'HTTP_AUTHORIZATION': f'Bearer {self.token}'}

    def _log_exercise(self, log_date, name, muscle_group, sets):
        payload = {
            'date': log_date.isoformat(),
            'notes': '',
            'exercises': [{
                'name': name,
                'muscle_group': muscle_group,
                'order': 0,
                'sets': [{'set_number': i + 1, 'weight_kg': w, 'reps': r} for i, (w, r) in enumerate(sets)],
            }],
        }
        resp = self.client.post(reverse('workout-logs'), payload, format='json', **self.auth_header())
        self.assertEqual(resp.status_code, 201)

    def test_comparison_computes_1rm_change_pct_vs_last_week(self):
        # Thursday of last calendar (Mon–Sun) week — safely inside "last
        # week" regardless of which weekday the test happens to run on.
        this_monday = self.today - timedelta(days=self.today.weekday())
        last_week_day = this_monday - timedelta(days=4)
        self._log_exercise(last_week_day, 'Biceps Curl', 'biceps', [(30, 6)])
        self._log_exercise(self.today, 'Biceps Curl', 'biceps', [(32, 6)])

        resp = self.client.get(
            reverse('workout-comparison'), {'end_date': self.today.isoformat()}, **self.auth_header()
        )
        self.assertEqual(resp.status_code, 200)
        exercises = resp.data['exercises']
        self.assertEqual(len(exercises), 1)
        curl = exercises[0]
        self.assertEqual(curl['name'], 'Biceps Curl')
        self.assertEqual(len(curl['sets']), 1)
        s = curl['sets'][0]
        # Epley: 1RM = weight * (1 + reps/30). Reps unchanged (6), so this
        # reduces to a plain weight ratio: 30->36.0, 32->38.4, +6.7%.
        self.assertEqual(s['previous_1rm_kg'], 36.0)
        self.assertEqual(s['current_1rm_kg'], 38.4)
        self.assertAlmostEqual(s['change_pct'], 6.7, delta=0.1)

    def test_comparison_uses_1rm_so_more_reps_at_same_weight_counts_as_smaller_gain(self):
        # Same weight, more reps this week. Epley values extra reps less
        # than extra weight, so this shows real (but modest) progress —
        # unlike a raw volume ratio, which would overstate it as +33%.
        this_monday = self.today - timedelta(days=self.today.weekday())
        last_week_day = this_monday - timedelta(days=4)
        self._log_exercise(last_week_day, 'Bicep Curl', 'biceps', [(25, 6)])
        self._log_exercise(self.today, 'Bicep Curl', 'biceps', [(25, 8)])

        resp = self.client.get(
            reverse('workout-comparison'), {'end_date': self.today.isoformat()}, **self.auth_header()
        )
        s = resp.data['exercises'][0]['sets'][0]
        self.assertEqual(s['previous_1rm_kg'], 30.0)
        self.assertAlmostEqual(s['current_1rm_kg'], 31.7, delta=0.05)
        # change_pct is derived from the rounded 1RM values above (31.7 vs
        # 30.0) so the percentage a user sees always matches what they'd get
        # hand-calculating from the two numbers shown on screen.
        self.assertAlmostEqual(s['change_pct'], 5.7, delta=0.1)

    def test_comparison_skips_exercises_without_a_prior_week_match(self):
        self._log_exercise(self.today, 'Brand New Exercise', 'chest', [(40, 8)])

        resp = self.client.get(
            reverse('workout-comparison'), {'end_date': self.today.isoformat()}, **self.auth_header()
        )
        self.assertEqual(resp.data['exercises'], [])

    def test_comparison_never_crosses_program_days_even_with_a_shared_exercise_name(self):
        # A 3-day full-body split where Day 1 and Day 3 both happen to
        # include "Push-ups" — they must be compared against their OWN
        # history only, never against each other.
        program_resp = self.client.post(reverse('user-programs'), {
            'name': 'Full Body', 'days': [
                {'name': 'Day 1', 'order': 0, 'exercises': [{'name': 'Push-ups', 'muscle_group': 'chest', 'order': 0}]},
                {'name': 'Day 2', 'order': 1, 'exercises': [{'name': 'Rows', 'muscle_group': 'back', 'order': 0}]},
                {'name': 'Day 3', 'order': 2, 'exercises': [{'name': 'Push-ups', 'muscle_group': 'chest', 'order': 0}]},
            ],
        }, format='json', **self.auth_header())
        day1_id = program_resp.data['days'][0]['id']
        day3_id = program_resp.data['days'][2]['id']

        this_monday = self.today - timedelta(days=self.today.weekday())
        last_monday = this_monday - timedelta(days=7)

        # Last week: only Day 3's Push-ups were logged, at a heavy load.
        self._log_with_program_day(last_monday, day3_id, 'Push-ups', 'chest', 0)  # bodyweight-style, reps only
        WorkoutSet.objects.filter(
            exercise_entry__workout_log__program_day_id=day3_id, exercise_entry__workout_log__date=last_monday
        ).update(reps=30)

        # This week: only Day 1's Push-ups logged, at a much lower rep count.
        self._log_with_program_day(this_monday, day1_id, 'Push-ups', 'chest', 0)
        WorkoutSet.objects.filter(
            exercise_entry__workout_log__program_day_id=day1_id, exercise_entry__workout_log__date=this_monday
        ).update(reps=10)

        resp = self.client.get(
            reverse('workout-comparison'), {'end_date': self.today.isoformat()}, **self.auth_header()
        )
        # Day 1's Push-ups has nothing to compare against (Day 1 wasn't done
        # last week) — it must NOT be silently matched to Day 3's data.
        self.assertEqual(resp.data['exercises'], [])

    def test_export_returns_valid_xlsx(self):
        self._log_exercise(self.today, 'Deadlift', 'back', [(100, 5), (110, 3)])

        resp = self.client.get(
            reverse('workout-export'), {'scope': 'week', 'end_date': self.today.isoformat()}, **self.auth_header()
        )
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(
            resp['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        from io import BytesIO
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(resp.content))
        ws = wb.active
        self.assertEqual(ws.cell(row=1, column=1).value, 'Mashq')
        self.assertTrue(ws.cell(row=1, column=2).value.startswith('Hafta 1'))
        # Two sets logged ⇒ two per-set sub-columns under that week.
        self.assertEqual(ws.cell(row=2, column=2).value, '1-padxod')
        self.assertEqual(ws.cell(row=2, column=3).value, '2-padxod')
        # No active program — the log falls into the "Erkin mashqlar" block.
        self.assertEqual(ws.cell(row=3, column=1).value, 'Erkin mashqlar')
        self.assertIn('Deadlift', ws.cell(row=4, column=1).value)
        self.assertEqual(ws.cell(row=4, column=2).value, '100.0kg×5')
        self.assertEqual(ws.cell(row=4, column=3).value, '110.0kg×3')

    def test_export_rows_are_program_days_and_exercises_columns_are_weeks(self):
        # Program: Day 1 (chest), Day 2 (back).
        program_resp = self.client.post(reverse('user-programs'), {
            'name': 'Split', 'days': [
                {'name': 'Day 1', 'order': 0, 'exercises': [{'name': 'Bench Press', 'muscle_group': 'chest', 'order': 0}]},
                {'name': 'Day 2', 'order': 1, 'exercises': [{'name': 'Deadlift', 'muscle_group': 'back', 'order': 0}]},
            ],
        }, format='json', **self.auth_header())
        program_id = program_resp.data['id']
        day1_id = program_resp.data['days'][0]['id']
        day2_id = program_resp.data['days'][1]['id']
        self.client.post(reverse('user-program-activate', args=[program_id]), **self.auth_header())

        this_monday = self.today - timedelta(days=self.today.weekday())
        last_monday = this_monday - timedelta(days=7)

        # Last week: did Day 1 then Day 2, in that (expected) order.
        self._log_with_program_day(last_monday, day1_id, 'Bench Press', 'chest', 60)
        self._log_with_program_day(last_monday + timedelta(days=3), day2_id, 'Deadlift', 'back', 100)
        # This week: skipped Day 1 entirely, only did Day 2 — it's the
        # week's *first* visit, but it must still land in Day 2's row.
        self._log_with_program_day(this_monday + timedelta(days=1), day2_id, 'Deadlift', 'back', 105)

        resp = self.client.get(
            reverse('workout-export'), {'scope': 'month', 'end_date': self.today.isoformat()}, **self.auth_header()
        )
        from io import BytesIO
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(resp.content))
        ws = wb.active

        # Rows: 1=week header, 2=set sub-header, 3=Day 1 block header,
        # 4=Bench Press, 5=spacer, 6=Day 2 block header, 7=Deadlift — in the
        # program's own order.
        self.assertEqual(ws.cell(row=3, column=1).value, 'Day 1')
        self.assertIn('Bench Press', ws.cell(row=4, column=1).value)
        self.assertEqual(ws.cell(row=6, column=1).value, 'Day 2')
        self.assertIn('Deadlift', ws.cell(row=7, column=1).value)

        # Only one set per session logged here, so each week is a single
        # sub-column: col 2 = last week ("Hafta 1"), col 3 = this week
        # ("Hafta 2"), each labeled "1-padxod" underneath.
        self.assertTrue(ws.cell(row=1, column=2).value.startswith('Hafta 1'))
        self.assertTrue(ws.cell(row=1, column=3).value.startswith('Hafta 2'))
        self.assertEqual(ws.cell(row=2, column=2).value, '1-padxod')
        self.assertEqual(ws.cell(row=2, column=3).value, '1-padxod')

        # Bench Press (Day 1) only happened last week.
        self.assertIn('60.0kg', ws.cell(row=4, column=2).value)
        self.assertIsNone(ws.cell(row=4, column=3).value)

        # Deadlift (Day 2) happened both weeks, at different weights — each
        # week's column must show THAT week's own session, never misfiled
        # into Day 1's row just because it was that week's first visit.
        self.assertIn('100.0kg', ws.cell(row=7, column=2).value)
        self.assertIn('105.0kg', ws.cell(row=7, column=3).value)

    def _log_with_program_day(self, log_date, program_day_id, name, muscle_group, weight_kg):
        payload = {
            'date': log_date.isoformat(),
            'notes': '',
            'program_day': program_day_id,
            'exercises': [{
                'name': name, 'muscle_group': muscle_group, 'order': 0,
                'sets': [{'set_number': 1, 'weight_kg': weight_kg, 'reps': 6}],
            }],
        }
        resp = self.client.post(reverse('workout-logs'), payload, format='json', **self.auth_header())
        self.assertEqual(resp.status_code, 201)


class UserProgramTests(BaseAPITestCase):
    def setUp(self):
        super().setUp()
        self.user = User.objects.create_user(email='planner@example.com', password='strongpass123')
        self.user.is_active = True
        self.user.save()
        resp = self.client.post(reverse('login'), {'email': 'planner@example.com', 'password': 'strongpass123'})
        self.token = resp.data['access']
        self.today = timezone.now().date()

    def auth_header(self):
        return {'HTTP_AUTHORIZATION': f'Bearer {self.token}'}

    def _create_program(self, name='My Split', days=None):
        payload = {
            'name': name,
            'days': days or [
                {'name': '1-kun (Ko\'krak)', 'order': 0, 'exercises': [
                    {'name': 'Bench Press', 'muscle_group': 'chest', 'order': 0},
                ]},
                {'name': '2-kun (Orqa)', 'order': 1, 'exercises': [
                    {'name': 'Deadlift', 'muscle_group': 'back', 'order': 0},
                ]},
            ],
        }
        return self.client.post(reverse('user-programs'), payload, format='json', **self.auth_header())

    def test_create_program_with_nested_days_and_exercises(self):
        resp = self._create_program()
        self.assertEqual(resp.status_code, 201)
        program = UserProgram.objects.get(user=self.user)
        self.assertEqual(program.days.count(), 2)
        self.assertEqual(program.days.first().exercises.first().name, 'Bench Press')

    def test_list_programs_scoped_to_user(self):
        self._create_program()
        other = User.objects.create_user(email='other_planner@example.com', password='strongpass123')
        UserProgram.objects.create(user=other, name='Someone else\'s plan')

        resp = self.client.get(reverse('user-programs'), **self.auth_header())
        self.assertEqual(len(resp.data), 1)

    def test_update_program_replaces_days(self):
        self._create_program()
        program_id = UserProgram.objects.get(user=self.user).id

        resp = self.client.patch(reverse('user-program-detail', args=[program_id]), {
            'days': [{'name': 'Full Body', 'order': 0, 'exercises': [
                {'name': 'Squat', 'muscle_group': 'legs', 'order': 0},
            ]}],
        }, format='json', **self.auth_header())
        self.assertEqual(resp.status_code, 200)

        program = UserProgram.objects.get(pk=program_id)
        self.assertEqual(program.days.count(), 1)
        self.assertEqual(program.days.first().name, 'Full Body')

    def test_editing_a_day_by_id_preserves_workout_log_links(self):
        self._create_program()
        program = UserProgram.objects.get(user=self.user)
        day1 = program.days.order_by('order').first()
        bench = day1.exercises.first()

        # A log was made against this exact day before any edit happens.
        log = WorkoutLog.objects.create(user=self.user, date=self.today, program_day=day1)

        # Rename the day and tweak its exercise, but keep both ids —
        # this must update in place, not delete-and-recreate.
        resp = self.client.patch(reverse('user-program-detail', args=[program.id]), {
            'days': [
                {'id': day1.id, 'name': 'Renamed Day', 'order': 0, 'exercises': [
                    {'id': bench.id, 'name': 'Bench Press', 'muscle_group': 'chest', 'order': 0},
                ]},
                {'name': '2-kun (Orqa)', 'order': 1, 'exercises': [
                    {'name': 'Deadlift', 'muscle_group': 'back', 'order': 0},
                ]},
            ],
        }, format='json', **self.auth_header())
        self.assertEqual(resp.status_code, 200)

        day1.refresh_from_db()
        self.assertEqual(day1.name, 'Renamed Day')
        log.refresh_from_db()
        self.assertEqual(log.program_day_id, day1.id)  # link survived the edit

    def test_removing_a_day_on_edit_still_nulls_out_its_logs(self):
        self._create_program()
        program = UserProgram.objects.get(user=self.user)
        day1, day2 = program.days.order_by('order')
        log = WorkoutLog.objects.create(user=self.user, date=self.today, program_day=day2)

        # New payload omits day2 entirely — it was genuinely deleted by the user.
        self.client.patch(reverse('user-program-detail', args=[program.id]), {
            'days': [{'id': day1.id, 'name': day1.name, 'order': 0, 'exercises': [
                {'name': 'Bench Press', 'muscle_group': 'chest', 'order': 0},
            ]}],
        }, format='json', **self.auth_header())

        self.assertFalse(UserProgramDay.objects.filter(id=day2.id).exists())
        log.refresh_from_db()
        self.assertIsNone(log.program_day_id)

    def test_delete_program(self):
        self._create_program()
        program_id = UserProgram.objects.get(user=self.user).id
        resp = self.client.delete(reverse('user-program-detail', args=[program_id]), **self.auth_header())
        self.assertEqual(resp.status_code, 204)
        self.assertFalse(UserProgram.objects.filter(id=program_id).exists())

    def test_cannot_access_other_users_program(self):
        other = User.objects.create_user(email='other_planner2@example.com', password='strongpass123')
        other_program = UserProgram.objects.create(user=other, name='Not yours')
        resp = self.client.patch(
            reverse('user-program-detail', args=[other_program.id]), {'name': 'Hijacked'},
            format='json', **self.auth_header()
        )
        self.assertEqual(resp.status_code, 404)

    def test_activate_deactivates_others(self):
        self._create_program('Plan A')
        self._create_program('Plan B')
        plan_a, plan_b = UserProgram.objects.filter(user=self.user).order_by('id')

        self.client.post(reverse('user-program-activate', args=[plan_a.id]), **self.auth_header())
        self.client.post(reverse('user-program-activate', args=[plan_b.id]), **self.auth_header())

        plan_a.refresh_from_db()
        plan_b.refresh_from_db()
        self.assertFalse(plan_a.is_active)
        self.assertTrue(plan_b.is_active)

    def test_active_endpoint(self):
        resp = self.client.get(reverse('user-program-active'), **self.auth_header())
        self.assertIsNone(resp.data)

        self._create_program('Plan A')
        program = UserProgram.objects.get(user=self.user)
        self.client.post(reverse('user-program-activate', args=[program.id]), **self.auth_header())

        resp = self.client.get(reverse('user-program-active'), **self.auth_header())
        self.assertEqual(resp.data['id'], program.id)

    def test_copy_from_catalog_creates_independent_snapshot(self):
        category = Category.objects.create(name_en='Chest', slug='chest')
        exercise = Exercise.objects.create(name_en='Bench Press', category=category)
        catalog_program = Program.objects.create(name_en='Push Pull Legs')
        catalog_day = ProgramDay.objects.create(program=catalog_program, name_en='Push Day', order=0)
        catalog_day.exercises.add(exercise)

        resp = self.client.post(
            reverse('user-program-copy'), {'program_id': catalog_program.id, 'lang': 'en'},
            format='json', **self.auth_header()
        )
        self.assertEqual(resp.status_code, 201)

        user_program = UserProgram.objects.get(user=self.user)
        self.assertEqual(user_program.name, 'Push Pull Legs')
        self.assertEqual(user_program.days.count(), 1)
        copied_exercise = user_program.days.first().exercises.first()
        self.assertEqual(copied_exercise.name, 'Bench Press')
        self.assertEqual(copied_exercise.muscle_group, '')  # left for the user to assign

        # Editing the catalog afterwards must never touch the user's copy.
        catalog_day.name_en = 'Renamed'
        catalog_day.save()
        user_program.refresh_from_db()
        self.assertEqual(user_program.days.first().name, 'Push Day')

    def test_workout_log_cannot_reference_another_users_program_day(self):
        other = User.objects.create_user(email='other_planner3@example.com', password='strongpass123')
        other_program = UserProgram.objects.create(user=other, name='Not yours')
        other_day = UserProgramDay.objects.create(program=other_program, name='Day 1')

        payload = {
            'date': self.today.isoformat(),
            'notes': '',
            'program_day': other_day.id,
            'exercises': [{'name': 'Bench Press', 'muscle_group': 'chest', 'order': 0, 'sets': []}],
        }
        resp = self.client.post(reverse('workout-logs'), payload, format='json', **self.auth_header())
        self.assertEqual(resp.status_code, 400)

    def test_workout_summary_includes_program_adherence(self):
        self._create_program()
        program = UserProgram.objects.get(user=self.user)
        day1, day2 = program.days.order_by('order')

        self.client.post(reverse('workout-logs'), {
            'date': self.today.isoformat(), 'notes': '', 'program_day': day1.id,
            'exercises': [{'name': 'Bench Press', 'muscle_group': 'chest', 'order': 0, 'sets': []}],
        }, format='json', **self.auth_header())
        self.client.post(reverse('user-program-activate', args=[program.id]), **self.auth_header())

        resp = self.client.get(reverse('workout-summary'), {'end_date': self.today.isoformat()}, **self.auth_header())
        adherence = resp.data['program_adherence']
        self.assertEqual(adherence['target_days'], 2)
        self.assertEqual(adherence['completed_days'], 1)

    def test_program_adherence_ignores_end_date_used_to_browse_a_past_week(self):
        """end_date navigates which week's muscle-group chart is shown, but
        program_adherence must always answer for the REAL current week —
        otherwise browsing history would show misleadingly stale progress."""
        self._create_program()
        program = UserProgram.objects.get(user=self.user)
        day1 = program.days.order_by('order').first()
        target_days = program.days.count()

        # Logged against Day 1 in the REAL current week.
        self.client.post(reverse('workout-logs'), {
            'date': self.today.isoformat(), 'notes': '', 'program_day': day1.id,
            'exercises': [{'name': 'Bench Press', 'muscle_group': 'chest', 'order': 0, 'sets': []}],
        }, format='json', **self.auth_header())
        self.client.post(reverse('user-program-activate', args=[program.id]), **self.auth_header())

        # Browse the chart to a week *before* this one, with no logs in it.
        last_week_day = self.today - timedelta(days=7)
        resp = self.client.get(
            reverse('workout-summary'),
            {'end_date': last_week_day.isoformat(), 'today': self.today.isoformat()},
            **self.auth_header(),
        )
        # The chart reflects the browsed (empty) past week...
        self.assertEqual(resp.data['total_workout_days'], 0)
        # ...but adherence still reports the real current week's progress —
        # if it had (incorrectly) followed end_date, this would read 0.
        self.assertEqual(resp.data['program_adherence']['completed_days'], 1)
        self.assertEqual(resp.data['program_adherence']['target_days'], target_days)
